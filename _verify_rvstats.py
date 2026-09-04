# -*- coding: utf-8 -*-
# 验证：复盘统计 4 卡片无徽章布局 + 同名项按位置分开储存
import time, re
from pathlib import Path
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

BASE = "http://127.0.0.1:8907"
SHOT = "_rv_dl/shot_rvstats.png"
REPO = Path(__file__).resolve().parent

ws = load_workbook(REPO / "review" / "2026-08.xlsx", read_only=True, data_only=True)["月度汇总"]
order, first_vals = [], None
for row in ws.iter_rows(min_col=2, max_col=34, values_only=True):
    item = row[0].strip() if isinstance(row[0], str) else None
    if item == "项目" or not item:
        continue
    order.append(item)
    has_num = any(isinstance(v, (int, float)) for v in row[1:])
    if first_vals is None and has_num:
        first_vals = (item, row[1], row[2])
print(f"expected: 61 行, 首项 {first_vals[0]}")

ok = True
def check(name, cond):
    global ok
    print(("PASS  " if cond else "FAIL  ") + name)
    if not cond:
        ok = False

def login(pg, user):
    pg.fill("#auth-username", user)
    pg.keyboard.press("Enter")
    pg.wait_for_function("()=>document.querySelector('#auth-overlay').hidden===true", timeout=20000)
    pg.click("#brand-btn")
    pg.wait_for_timeout(400)
    pg.click("#tab-review")
    pg.wait_for_function("()=>(!document.querySelector('#rvw-cards').hidden&&document.querySelectorAll('#rvw-cards .rvw-card').length>0)||!document.querySelector('#rvw-empty').hidden", timeout=40000)

def cards_info(pg):
    return pg.evaluate("""()=>[...document.querySelectorAll('#rvw-cards .rvw-card')].map(c=>{
      const tbl=c.querySelector('table');
      const rows=[...tbl.querySelectorAll('tr')];
      return {t:c.querySelector('.rvw-card-head').textContent,
              n:rows.length,
              items:rows.slice(1).map(r=>r.querySelector('td.rvw-item').textContent)};})""")

def cell(pg, ci, ri, tdidx):
    return pg.evaluate(f"""()=>{{
      const c=document.querySelectorAll('#rvw-cards .rvw-card')[{ci}];
      const tr=c.querySelectorAll('table tr')[{ri}];
      return tr.querySelectorAll('td')[{tdidx}].textContent;}}""")

with sync_playwright() as p:
    b = p.chromium.launch(channel="msedge", headless=True)
    pg = b.new_page(viewport={"width": 1440, "height": 1000})
    errors = []
    pg.on("pageerror", lambda e: errors.append(str(e)))
    pg.goto(BASE + "/?v=" + str(int(time.time() * 1000)), wait_until="domcontentloaded")
    login(pg, "songningning")

    info = cards_info(pg)
    ts = [c["t"] for c in info]
    check(f"5 卡（4 模板卡+其他数据） {ts}", ts == ["电量与结算概览", "交易电费明细", "交易电费明细", "策略影响电价", "其他数据"])
    check("模板卡全空（无回退错显，标题行为空串）", pg.evaluate("""()=>{
      const c=document.querySelectorAll('#rvw-cards .rvw-card')[0];
      return [...c.querySelectorAll('table tr')].slice(1).every(r=>[...r.querySelectorAll('td')].slice(1).every(td=>r.classList.contains('rvw-hdr')?td.textContent==='':td.textContent==='—'));}"""))
    check("其他数据卡=测试上传的 2 项", info[4]["items"] == ["售电收入", "购电成本"])

    pg.select_option("#rvw-month", "2026-08")
    pg.wait_for_function("()=>document.querySelector('#rvw-count').textContent.includes('共 54 项')&&document.querySelectorAll('#rvw-cards .rvw-card')[0].querySelectorAll('table tr').length===15", timeout=40000)
    info = cards_info(pg)
    check(f"卡1 {info[0]['n']} 行(表头+14)", info[0]["n"] == 15 and info[0]["items"] == order[:14])
    check(f"卡2 {info[1]['n']} 行(表头+21)", info[1]["n"] == 22 and info[1]["items"] == order[14:35])
    check(f"卡3 {info[2]['n']} 行(表头+8, 首行交易电费（元）)", info[2]["n"] == 9 and info[2]["items"] == [order[14]] + order[35:42])
    check(f"卡4 {info[3]['n']} 行(表头+19)", info[3]["n"] == 20 and info[3]["items"] == order[42:61])
    check("卡1 首行月汇总/日1", cell(pg, 0, 1, 1) == "13998.236" and cell(pg, 0, 1, 2) == "437.192")
    check("小数位：卡1 结算电价 3 位/结算电费 2 位", cell(pg, 0, 4, 2) == "372.000" and cell(pg, 0, 5, 2) == "162635.42")
    def cls(pg, ci, ri, tdidx):
        return pg.evaluate(f"""()=>{{
          const c=document.querySelectorAll('#rvw-cards .rvw-card')[{ci}];
          const tr=c.querySelectorAll('table tr')[{ri}];
          return tr.querySelectorAll('td')[{tdidx}].className;}}""")
    check("正负着色：卡2 申报策略影响电费 负=蓝 / 交易电费（元）+若全量进日前 不变色 / 零值默认 / 卡4 默认 / 卡1 不着色",
          "rvw-neg" in cls(pg, 1, 5, 1) and "rvw-pos" not in cls(pg, 1, 1, 1) and "rvw-pos" not in cls(pg, 1, 2, 1)
          and "rvw-pos" not in cls(pg, 1, 3, 1) and "rvw-neg" not in cls(pg, 1, 3, 1)
          and "rvw-pos" not in cls(pg, 3, 1, 2) and "rvw-neg" not in cls(pg, 3, 1, 2)
          and "rvw-pos" not in cls(pg, 0, 1, 2) and "rvw-neg" not in cls(pg, 0, 1, 2))
    dup = "预测偏差产生电费"
    d2 = [i for i, it in enumerate(info[1]["items"]) if it == dup][0]
    d3 = [i for i, it in enumerate(info[2]["items"]) if it == dup]
    check(f"卡2 重复项回退显示合并值 {cell(pg, 1, d2+1, 2)}", cell(pg, 1, d2+1, 2) == "0.00")
    check(f"卡3 重复项两处回退显示同值 {[cell(pg, 2, d+1, 2) for d in d3]}", all(cell(pg, 2, d+1, 2) == "0.00" for d in d3))

    pg.hover("#rvw-cards .rvw-card:nth-child(1) table tr:nth-child(3) td:nth-child(5)")
    hl = pg.evaluate("""()=>{const c=document.querySelector('#rvw-cards .rvw-card:nth-child(1) .rvw-hlc');return c?{ri:[...c.parentElement.parentElement.children].indexOf(c.parentElement)+1,n:document.querySelectorAll('#rvw-cards .rvw-hlc').length}:{ri:-1,n:0};}""")
    check(f"卡内悬停列高亮+焦点格 (row={hl['ri']} 全局焦点数={hl['n']})", hl["n"] == 1 and hl["ri"] == 3)
    pg.hover("#rvw-cards .rvw-card:nth-child(1) table tr:nth-child(4) td:nth-child(5)")
    hl = pg.evaluate("""()=>{const c=document.querySelector('#rvw-cards .rvw-card:nth-child(1) .rvw-hlc');return c?[...c.parentElement.parentElement.children].indexOf(c.parentElement)+1:-1;}""")
    check(f"同列下移焦点格跟随 (row={hl})", hl == 4)
    pg.mouse.move(5, 5)

    cnt = pg.locator("#rvw-count").text_content()
    check(f"计数 ({cnt})", "共 54 项" in (cnt or ""))
    check("导出 CSV 按钮不存在", pg.evaluate("()=>!document.querySelector('#rvw-export')"))
    pg.click("#rvw-refresh")
    pg.wait_for_function("()=>document.querySelectorAll('#rvw-cards .rvw-card')[0].querySelectorAll('table tr').length===15", timeout=40000)
    check("刷新后卡1仍 15 行", True)
    pg.screenshot(path=SHOT)

    pg2 = b.new_page(viewport={"width": 1440, "height": 900})
    pg2.goto(BASE + "/?v=" + str(int(time.time() * 1000)), wait_until="domcontentloaded")
    pg2.fill("#auth-username", "wanliyang")
    pg2.keyboard.press("Enter")
    pg2.wait_for_function("()=>document.querySelector('#auth-overlay').hidden===true", timeout=20000)
    pg2.click("#brand-btn")
    pg2.wait_for_timeout(400)
    pg2.click("#tab-review")
    pg2.wait_for_function("()=>!document.querySelector('#rvw-empty').hidden", timeout=40000)
    check("wanliyang 空态", pg2.locator("#rvw-empty").is_visible())

    # gly 2026-09（legacy 真实数据）：4 卡行数分布
    pg3 = b.new_page(viewport={"width": 1440, "height": 900})
    pg3.goto(BASE + "/?v=" + str(int(time.time() * 1000)), wait_until="domcontentloaded")
    pg3.fill("#auth-username", "gly")
    pg3.keyboard.press("Enter")
    pg3.wait_for_function("()=>document.querySelector('#auth-overlay').hidden===true", timeout=20000)
    pg3.click("#brand-btn")
    pg3.wait_for_timeout(400)
    pg3.click("#tab-review")
    pg3.wait_for_function("()=>document.querySelectorAll('#rvw-cards .rvw-card').length===4", timeout=40000)
    info3 = cards_info(pg3)
    check(f"gly 2026-09 卡行数 {[c['n'] for c in info3]}", [c["n"] for c in info3] == [15, 22, 9, 20])
    cnt2 = pg3.locator("#rvw-count").text_content()
    check(f"gly 计数含更新时间 ({cnt2})", "共 54 项" in (cnt2 or "") and "更新于" in (cnt2 or ""))

    # 全员总览页（4 卡片）
    pg3.click("#tab-allsum")
    pg3.wait_for_function("()=>!document.querySelector('#allsum-cards').hidden&&document.querySelectorAll('#allsum-cards .rvw-card').length===4", timeout=40000)
    a0 = pg3.evaluate("""()=>[...document.querySelectorAll('#allsum-cards .rvw-card')].map(c=>{
      const t=c.querySelector('table');const rows=[...t.rows];
      return {t:c.querySelector('.rvw-card-head').textContent,n:rows.length,
        heads:[...rows[0].children].map(th=>th.textContent),
        items:rows.slice(1).map(r=>r.children[0].textContent)};})""")
    tsA = [c["t"] for c in a0]
    check(f"全员总览 4 卡片 {tsA}", tsA == ["电量与结算概览", "交易电费明细", "交易电费明细", "策略影响电价"])
    check(f"全员总览卡行数 {[c['n'] for c in a0]} (表头+14/21/8/19)", [c["n"] for c in a0] == [15, 22, 9, 20])
    heads0 = a0[0]["heads"]
    check(f"账号表头 12 个不含管理员 (首={heads0[0]})", heads0[0] == "项目" and len(heads0) == 13 and "宋宁宁" in heads0 and "管理员" not in heads0)
    check(f"全员总览计数 ({a0 and pg3.locator('#allsum-count').text_content()})", "61 项 × 12 个账号" in (pg3.locator("#allsum-count").text_content() or "") and "有数据 宋宁宁" in (pg3.locator("#allsum-count").text_content() or ""))
    pg3.select_option("#allsum-month", "2026-08")
    pg3.wait_for_function("()=>document.querySelector('#allsum-count').textContent.includes('· 2026-08 ·')", timeout=40000)
    a1 = pg3.evaluate("""()=>{const c0=document.querySelectorAll('#allsum-cards .rvw-card')[0].querySelector('table');
      const find=(card,item)=>{const r=[...card.rows].find(r=>r.children[0].textContent===item);return r?[...r.children].map(td=>td.textContent):null;};
      const nDup=[...document.querySelectorAll('#allsum-cards .rvw-card table')].reduce((n,t)=>n+[...t.rows].filter(r=>r.children[0].textContent==='预测偏差产生电费').length,0);
      const card2=document.querySelectorAll('#allsum-cards .rvw-card')[1].querySelector('table');
      const dup2=find(card2,'预测偏差产生电费');
      return {qty:find(c0,'实际用电量（MWh）').slice(0,3),dup2:dup2.slice(0,3),nDup:nDup};}""")
    check(f"2026-08 宋宁宁列月汇总值 (首行={a1['qty'][2]})", a1["qty"][0] == "实际用电量（MWh）" and a1["qty"][2] == "13998.236" and a1["qty"][1] == "—")
    check(f"重复项 3 处同显回退合并值 ({a1['dup2'][2]}/共{a1['nDup']}处)", a1["dup2"][2] == "0.00" and a1["nDup"] == 3)
    acls = pg3.evaluate("""()=>{const c2=document.querySelectorAll('#allsum-cards .rvw-card')[1].querySelector('table');
      const fee=[...c2.rows].find(r=>r.children[0].textContent==='交易电费（元）').children[2];
      const chg=[...c2.rows].find(r=>r.children[0].textContent==='日前策略导致电费变动').children[2];
      const neg=[...c2.rows].find(r=>r.children[0].textContent==='申报策略影响电费').children[2];
      const c1=document.querySelectorAll('#allsum-cards .rvw-card')[0].querySelector('table');
      const qty=[...c1.rows].find(r=>r.children[0].textContent==='实际用电量（MWh）').children[2];
      return {fee:fee.className,qty:qty.className,chg:chg.className,neg:neg.className};}""")
    check("全员总览正负着色：申报策略影响电费 负=蓝 / 交易电费（元）不变色 / 电量不着色", "rvw-neg" in acls["neg"] and "rvw-pos" not in acls["fee"] and "rvw-pos" not in acls["qty"] and "rvw-neg" not in acls["qty"])
    pg3.click("#tab-review")
    pg3.wait_for_function("()=>!document.querySelector('#rvw-cards').hidden", timeout=40000)
    check("切回复盘统计正常", True)

    check("无 JS 错误", not errors)
    if errors: print("JS errors:", errors[:3])
    b.close()

# ============ 分开储存回归：gly 2026-07 快照→写入不同值→页面验证→try/finally 还原 ============
import psycopg2
password = re.search(r"password=(.+)", (Path.home() / ".workbuddy" / "db_secret.txt").read_text(encoding="utf-8")).group(1).strip()

def db():
    c = psycopg2.connect(host="aws-1-ap-northeast-1.pooler.supabase.com", port=5432,
                         user="postgres.kputrrbbcmxbnhwhxvoo", password=password, dbname="postgres",
                         connect_timeout=30, keepalives=1, keepalives_idle=5, keepalives_interval=2)
    c.autocommit = True
    return c

conn = db(); cur = conn.cursor()
cur.execute("select item,day,value,seq,ord from review_summary where username='gly' and month='2026-07'")
snap = cur.fetchall()
print(f"snapshot gly 2026-07 rows: {len(snap)}")
if any(r[0] == "预测偏差产生电费" for r in snap):
    print("WARN: snapshot polluted by synthetic rows; abort without writes"); conn.close(); raise SystemExit(2)
try:
    ITEM = "预测偏差产生电费"
    for seq_v, val in ((17, 111.111), (37, 222.222), (40, 333.333)):
        cur.execute("delete from review_summary where username='gly' and month='2026-07' and seq=%s", (seq_v,))
        for d in range(32):
            cur.execute("insert into review_summary (username,month,item,day,value,seq) values ('gly','2026-07',%s,%s,%s,%s)", (ITEM, d, val, seq_v))
    print("synthetic rows written")
    with sync_playwright() as p:
        b = p.chromium.launch(channel="msedge", headless=True)
        pg = b.new_page(viewport={"width": 1440, "height": 1000})
        errors = []
        pg.on("pageerror", lambda e: errors.append(str(e)))
        pg.goto(BASE + "/?v=" + str(int(time.time() * 1000)), wait_until="domcontentloaded")
        login(pg, "gly")
        pg.select_option("#rvw-month", "2026-07")
        pg.wait_for_function("()=>document.querySelector('#rvw-count').textContent.includes('共 5 项')&&document.querySelectorAll('#rvw-cards .rvw-card')[1].querySelectorAll('table tr').length===22", timeout=40000)
        v17 = cell(pg, 1, 4, 2)
        v37 = cell(pg, 2, 4, 2)
        v40 = cell(pg, 2, 7, 2)
        check(f"分开储存：卡2 预测偏差={v17} 卡3 两处={v37}/{v40}", v17 == "111.11" and v37 == "222.22" and v40 == "333.33")
        check("无 JS 错误", not errors)
        b.close()
finally:
    try:
        cur.execute("delete from review_summary where username='gly' and month='2026-07'")
        for item, day, value, seq_v, ord_v in snap:
            cur.execute("insert into review_summary (username,month,item,day,value,seq,ord) values ('gly','2026-07',%s,%s,%s,%s,%s)", (item, day, value, seq_v, ord_v))
    except Exception as e:
        print("main conn restore failed:", e)
        for attempt in range(3):
            try:
                c2 = db(); k = c2.cursor()
                k.execute("delete from review_summary where username='gly' and month='2026-07'")
                for item, day, value, seq_v, ord_v in snap:
                    k.execute("insert into review_summary (username,month,item,day,value,seq,ord) values ('gly','2026-07',%s,%s,%s,%s,%s)", (item, day, value, seq_v, ord_v))
                c2.close(); print("restored via fresh conn"); break
            except Exception as e2:
                print("fresh conn attempt", attempt, "failed:", e2)
    try:
        cur.execute("select count(*) from review_summary where username='gly' and month='2026-07'")
        print("restored rows:", cur.fetchone()[0], "(snapshot was", len(snap), ")")
    except Exception:
        c3 = db(); k = c3.cursor()
        k.execute("select count(*) from review_summary where username='gly' and month='2026-07'")
        print("restored rows:", k.fetchone()[0], "(snapshot was", len(snap), "via fresh conn)")
        c3.close()
    try: cur.close()
    except Exception: pass
    conn.close()
print("RESULT:", "PASS" if ok else "FAIL")
