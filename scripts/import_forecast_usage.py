"""导入「预测用电量」Excel 到 daily_curves.predicted_value（2026-09 批量）。

源文件：任意 xlsx，列结构 A=日期 B=时间("HH:00-HH+1:00") C=预估用电量
主键：(date, hour)，UPSERT + COALESCE 空值保护（Excel 空值不覆盖 DB 已有值）。
精度：电量 3 位小数。
用法：python import_forecast_usage.py <xlsx路径> [日期下限(可选, 默认全部)]
"""
import sys
import re
import psycopg2
from openpyxl import load_workbook

DB_HOST = "aws-1-ap-northeast-1.pooler.supabase.com"
DB_PORT = 5432
DB_NAME = "postgres"
DB_USER = "postgres.kputrrbbcmxbnhwhxvoo"
SECRET_FILE = r"C:/Users/ningning2.song/.workbuddy/db_secret.txt"


def load_password():
    txt = open(SECRET_FILE, encoding="utf-8").read()
    m = re.search(r"password=([A-Za-z0-9*#%]+)", txt)
    if not m:
        raise SystemExit("未在 db_secret.txt 中找到 password= 项")
    return m.group(1)


def parse_file(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    records = []
    seen = set()
    for r in rows:
        d, t, v = r[0], r[1], r[2]
        if d is None:
            raise ValueError("日期为空")
        ds = d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d)
        if t is None:
            raise ValueError(f"{ds}: 时间为空")
        hm = str(t).strip().split("-")[0].split(":")[0]
        try:
            hour = int(hm)
        except ValueError:
            raise ValueError(f"{ds} {t}: 时段无法解析")
        if not (0 <= hour <= 23):
            raise ValueError(f"{ds} {t}: 时段 {hour} 越界")
        if v is None or str(v).strip() == "":
            raise ValueError(f"{ds} {t}: 预测用电量为空")
        value = round(float(v), 3)
        key = (ds, hour)
        if key in seen:
            raise ValueError(f"{ds} {hour}:00 重复")
        seen.add(key)
        records.append((ds, hour, value))
    return records


def main():
    if len(sys.argv) < 2:
        raise SystemExit(__doc__)
    path = sys.argv[1]
    min_date = sys.argv[2] if len(sys.argv) > 2 else None
    records = parse_file(path)
    if min_date:
        records = [r for r in records if r[0] >= min_date]
    if not records:
        raise SystemExit("没有可导入记录")
    dates = sorted({r[0] for r in records})
    print(f"解析成功：{len(records)} 条，日期 {dates[0]} ~ {dates[-1]}，共 {len(dates)} 天")

    conn = psycopg2.connect(
        host=DB_HOST, port=DB_PORT, dbname=DB_NAME,
        user=DB_USER, password=load_password(), connect_timeout=15,
    )
    conn.autocommit = True
    cur = conn.cursor()
    cur.execute("SELECT MIN(date), MAX(date) FROM daily_curves")
    dmin, dmax = cur.fetchone()
    if min(dates) < str(dmin) or max(dates) > str(dmax):
        conn.close()
        raise SystemExit(f"日期越界：Excel {dates[0]}~{dates[-1]}，表范围 {dmin}~{dmax}")

    upsert = (
        "INSERT INTO daily_curves (date, hour, predicted_value) VALUES (%s, %s, %s) "
        "ON CONFLICT (date, hour) DO UPDATE SET "
        "predicted_value = COALESCE(EXCLUDED.predicted_value, daily_curves.predicted_value)"
    )
    by_date = {}
    for r in records:
        by_date.setdefault(r[0], []).append(r)
    total = 0
    for ds in sorted(by_date):
        batch = by_date[ds]
        cur.executemany(upsert, [(r[0], r[1], r[2]) for r in batch])
        total += cur.rowcount
        print(f"  {ds}: {len(batch)} 行")
    print(f"UPSERT 完成：{total} 行")

    # 复核
    cur.execute(
        "SELECT date, COUNT(*), COUNT(predicted_value) FROM daily_curves "
        "WHERE date BETWEEN %s AND %s GROUP BY date ORDER BY date",
        (dates[0], dates[-1]),
    )
    bad = 0
    for d, total, nonnull in cur.fetchall():
        mark = "OK" if nonnull == 24 else "!!"
        if nonnull != 24:
            bad += 1
        print(f"  {d}  total={total} 预测非空={nonnull} {mark}")
    cur.close()
    conn.close()
    if bad:
        raise SystemExit(f"有 {bad} 天预测不完整")
    print("全部日期预测 24/24 非空，导入成功")


if __name__ == "__main__":
    main()
