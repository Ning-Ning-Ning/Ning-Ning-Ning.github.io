#!/usr/bin/env python3
"""Validate and import actual, predicted, and adjusted electricity usage."""

from __future__ import annotations

import argparse
import json
import os
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from openpyxl import load_workbook

EXPECTED_HEADERS = (
    "年月日",
    "小时",
    "实际用电量MWh",
    "预测用电量MWh",
    "调整后用电量MWh",
)
ACTUAL_NULL_FROM = date(2026, 5, 24)
EXPECTED_HOURS = set(range(1, 25))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("excel", type=Path, help="Source Excel workbook")
    parser.add_argument("--url", default=os.getenv("SUPABASE_URL"), help="Supabase project URL")
    parser.add_argument(
        "--key",
        default=os.getenv("SUPABASE_IMPORT_KEY"),
        help="Temporary key allowed to write daily_curves; prefer SUPABASE_IMPORT_KEY",
    )
    parser.add_argument("--batch-size", type=int, default=500)
    parser.add_argument("--validate-only", action="store_true")
    return parser.parse_args()


def normalize_date(value: object, row_number: int) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    try:
        return datetime.fromisoformat(str(value)).date().isoformat()
    except ValueError as exc:
        raise ValueError(f"第 {row_number} 行日期无效：{value!r}") from exc


def optional_number(value: object, label: str, row_number: int) -> float | None:
    if value is None or value == "":
        return None
    number = float(value)
    if number < 0:
        raise ValueError(f"第 {row_number} 行{label}小于 0：{value!r}")
    return number


def read_records(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Sheet1"]
    rows = sheet.iter_rows(values_only=True)
    headers = tuple(next(rows))
    if headers != EXPECTED_HEADERS:
        raise ValueError(f"表头应为 {EXPECTED_HEADERS}，实际为 {headers}")

    records: list[dict[str, object]] = []
    per_date: Counter[str] = Counter()
    seen: set[tuple[str, int]] = set()
    predicted_count = 0
    adjusted_count = 0
    actual_count = 0

    for row_number, row in enumerate(rows, start=2):
        if len(row) != 5:
            raise ValueError(f"第 {row_number} 行字段数不是 5：{row!r}")

        raw_date, raw_hour, raw_actual, raw_predicted, raw_adjusted = row
        if raw_date is None or raw_hour is None or raw_actual is None:
            raise ValueError(f"第 {row_number} 行日期、小时或实际用电量为空：{row!r}")

        day = normalize_date(raw_date, row_number)
        excel_hour = int(raw_hour)
        if excel_hour not in EXPECTED_HOURS:
            raise ValueError(f"第 {row_number} 行小时不在 1..24：{raw_hour!r}")

        actual = optional_number(raw_actual, "实际用电量", row_number)
        predicted = optional_number(raw_predicted, "预测用电量", row_number)
        adjusted = optional_number(raw_adjusted, "调整后用电量", row_number)
        assert actual is not None
        if date.fromisoformat(day) >= ACTUAL_NULL_FROM:
            actual = None

        actual_count += actual is not None
        predicted_count += predicted is not None
        adjusted_count += adjusted is not None

        db_hour = excel_hour - 1
        key = (day, db_hour)
        if key in seen:
            raise ValueError(f"发现重复日期和时段：{key}")
        seen.add(key)
        per_date[day] += 1
        records.append(
            {
                "date": day,
                "hour": db_hour,
                "actual_value": actual,
                "predicted_value": predicted,
                "adjusted_value": adjusted,
            }
        )

    invalid_days = [day for day, count in per_date.items() if count != 24]
    if invalid_days:
        raise ValueError(f"以下日期不是 24 条记录：{invalid_days[:10]}")

    expected_count = len(per_date) * 24
    if len(records) != expected_count:
        raise ValueError(f"记录数不一致：{len(records)} != {expected_count}")

    print(
        f"校验完成：{len(records)} 条，{len(per_date)} 天，"
        f"{min(per_date)} 至 {max(per_date)}；实际 {actual_count} 个，"
        f"预测 {predicted_count} 个，原始调整后 {adjusted_count} 个。"
    )
    return records


def upload(records: list[dict[str, object]], url: str, key: str, batch_size: int) -> None:
    endpoint = f"{url.rstrip('/')}/rest/v1/daily_curves?on_conflict=date,hour"
    total_batches = (len(records) + batch_size - 1) // batch_size

    for batch_number, offset in enumerate(range(0, len(records), batch_size), start=1):
        payload = json.dumps(records[offset : offset + batch_size]).encode("utf-8")
        request = Request(
            endpoint,
            data=payload,
            method="POST",
            headers={
                "apikey": key,
                "Content-Type": "application/json",
                "Prefer": "resolution=merge-duplicates,return=minimal",
            },
        )
        try:
            with urlopen(request, timeout=60) as response:
                if response.status not in (200, 201, 204):
                    raise RuntimeError(f"Supabase 返回 HTTP {response.status}")
        except HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"第 {batch_number} 批导入失败：HTTP {exc.code} {detail}") from exc
        except URLError as exc:
            raise RuntimeError(f"第 {batch_number} 批网络错误：{exc.reason}") from exc

        print(f"已导入 {batch_number}/{total_batches} 批")


def main() -> int:
    args = parse_args()
    try:
        records = read_records(args.excel)
        if args.validate_only:
            return 0
        if not args.url or not args.key:
            raise ValueError("导入需要 SUPABASE_URL 和 SUPABASE_IMPORT_KEY")
        upload(records, args.url, args.key, args.batch_size)
    except (OSError, ValueError, RuntimeError) as exc:
        print(f"错误：{exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
